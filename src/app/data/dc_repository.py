"""Depot data loader with database-first approach, falling back to Excel file."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook

from ..config import settings
from ..models.domain import Depot
from ..db.supabase import get_supabase_client


def _normalize_dc_name(name: str) -> str:
    return name.strip()


def _load_depots_from_database() -> tuple[Depot, ...] | None:
    """Load depots from Supabase database. Returns None if database not available or empty."""
    supabase = get_supabase_client()
    if not supabase:
        return None
    
    try:
        response = supabase.table("depots").select("*").execute()
        if not response.data or len(response.data) == 0:
            return None
        
        depots: list[Depot] = []
        for row in response.data:
            try:
                # Handle date conversion if needed
                effective_date = row.get("effective_date")
                if effective_date and isinstance(effective_date, str):
                    from datetime import date
                    effective_date = date.fromisoformat(effective_date)
                
                depots.append(
                    Depot(
                        code=_normalize_dc_name(str(row["code"])),
                        latitude=float(row["latitude"]),
                        longitude=float(row["longitude"]),
                        effective_date=effective_date,
                    )
                )
            except (KeyError, ValueError, TypeError) as e:
                # Skip invalid rows but continue processing
                import logging
                logging.warning(f"Skipping invalid depot row: {e}")
                continue
        
        return tuple(depots) if depots else None
    except Exception as e:
        # If database query fails, return None to fall back to file
        import logging
        logging.debug(f"Database query failed, falling back to file: {e}")
        return None


def _load_depots_from_file(source: Path | None = None) -> tuple[Depot, ...]:
    """Load depots from Excel file."""
    workbook_path = (source or settings.dc_locations_file)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Depot workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    sheet = wb.active
    rows = sheet.iter_rows(min_row=1, values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError(f"Depot workbook '{workbook_path}' is empty.")

    header_map = {name: idx for idx, name in enumerate(header)}
    missing_columns = {"DC", "Latitude", "Longitude"} - set(header_map)
    if missing_columns:
        raise ValueError(f"Depot workbook missing columns: {', '.join(sorted(missing_columns))}")

    depots: list[Depot] = []
    for row in rows:
        dc_value = row[header_map["DC"]]
        lat_value = row[header_map["Latitude"]]
        lon_value = row[header_map["Longitude"]]
        if not dc_value:
            continue
        depots.append(
            Depot(
                code=_normalize_dc_name(str(dc_value)),
                latitude=float(lat_value),
                longitude=float(lon_value),
            )
        )
    return tuple(depots)


def _sync_depots_to_database(depots: tuple[Depot, ...]) -> None:
    """Sync depots from file to database. Only inserts if not already present."""
    supabase = get_supabase_client()
    if not supabase:
        return  # Database not configured, skip sync
    
    try:
        # Get existing depot codes
        existing_response = supabase.table("depots").select("code").execute()
        existing_codes = {row["code"].upper() for row in existing_response.data} if existing_response.data else set()
        
        # Insert only new depots
        new_depots = []
        for depot in depots:
            if depot.code.upper() not in existing_codes:
                new_depots.append({
                    "code": depot.code,
                    "name": depot.code,  # Use code as name if not provided
                    "latitude": depot.latitude,
                    "longitude": depot.longitude,
                    "effective_date": depot.effective_date.isoformat() if depot.effective_date else None,
                })
        
        if new_depots:
            supabase.table("depots").insert(new_depots).execute()
            # Clear cache so next call will load from database
            from .customers_repository import clear_dc_lookup_cache
            clear_dc_lookup_cache()
    except Exception as e:
        # If sync fails, continue without database - file-based fallback will work
        import logging
        logging.debug(f"Failed to sync depots to database (non-critical): {e}")
        pass


def get_depots(source: Path | None = None) -> tuple[Depot, ...]:
    """Get depots from database first, fall back to Excel file if needed.
    
    If database is empty or not configured, loads from Excel file and syncs to database.
    """
    # Try database first
    db_depots = _load_depots_from_database()
    if db_depots:
        return db_depots
    
    # Fall back to file
    file_depots = _load_depots_from_file(source)
    
    # Sync to database for next time (non-blocking)
    if file_depots:
        _sync_depots_to_database(file_depots)
    
    return file_depots
