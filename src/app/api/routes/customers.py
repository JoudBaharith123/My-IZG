"""Customer dataset endpoints."""

from __future__ import annotations

from typing import List
from datetime import datetime, timezone
from pathlib import Path
from io import BytesIO
import csv
import json

from fastapi import APIRouter, Query, status, UploadFile, File, Form, HTTPException
from openpyxl import load_workbook

from ...schemas.customers import (
    CitySummaryModel,
    CustomerLocationModel,
    CustomerLocationsResponse,
    CustomerStatsResponse,
    CustomerValidationResponse,
    ZoneSummaryModel,
)
from ...services.customers import (
    analyze_customer_issues,
    compute_customer_stats,
    compute_zone_summaries,
    list_customer_cities,
    list_customer_locations,
)
from ...data.customers_repository import set_active_customer_file, load_customers
from ...config import settings

router = APIRouter(prefix="/customers", tags=["customers"])


@router.get("/stats", response_model=CustomerStatsResponse, status_code=status.HTTP_200_OK)
def get_customer_stats() -> CustomerStatsResponse:
    return CustomerStatsResponse(**compute_customer_stats())


@router.get("/cities", response_model=List[CitySummaryModel], status_code=status.HTTP_200_OK)
def list_cities(limit: int | None = Query(default=None, ge=1, le=500)) -> List[CitySummaryModel]:
    city_list = list_customer_cities(limit=limit)
    return [CitySummaryModel(**entry) for entry in city_list]


@router.get("/zones", response_model=List[ZoneSummaryModel], status_code=status.HTTP_200_OK)
def list_customer_zones(city: str | None = Query(default=None, description="Optional city filter")) -> List[ZoneSummaryModel]:
    return [ZoneSummaryModel(**entry) for entry in compute_zone_summaries(city)]


@router.get("/locations", response_model=CustomerLocationsResponse, status_code=status.HTTP_200_OK)
def get_customer_locations(
    city: str | None = Query(default=None, description="Optional city or area filter"),
    zone: str | None = Query(default=None, description="Optional zone filter"),
    status: str | None = Query(default=None, description="Optional status filter"),
    agent_id: str | None = Query(default=None, description="Optional agent ID filter"),
    agent_name: str | None = Query(default=None, description="Optional agent name filter"),
    page: int = Query(default=1, ge=1, description="1-based page index for pagination"),
    page_size: int = Query(default=1000, ge=100, le=5000, description="Maximum number of records per page"),
    limit: int | None = Query(
        default=None,
        ge=1,
        le=10_000,
        description="Optional hard cap on returned records (overrides page_size when smaller).",
    ),
) -> CustomerLocationsResponse:
    effective_page_size = min(page_size, limit) if limit else page_size
    offset = (page - 1) * effective_page_size

    # Build filters dictionary (using lowercase with underscores for consistency)
    filters = {}
    if status:
        filters['status'] = status
    if agent_id:
        filters['agent_id'] = agent_id
    if agent_name:
        filters['agent_name'] = agent_name

    items, total = list_customer_locations(
        city=city,
        zone=zone,
        filters=filters if filters else None,
        offset=offset,
        limit=effective_page_size
    )
    has_next_page = (offset + len(items)) < total
    return CustomerLocationsResponse(
        items=[CustomerLocationModel(**record) for record in items],
        page=page,
        page_size=effective_page_size,
        total=total,
        has_next_page=has_next_page,
    )


@router.get("/validation", response_model=CustomerValidationResponse, status_code=status.HTTP_200_OK)
def validate_customer_dataset() -> CustomerValidationResponse:
    return CustomerValidationResponse.model_validate(analyze_customer_issues())


@router.post("/upload/preview")
async def preview_customer_file(file: UploadFile = File(...)) -> dict:
    """Preview CSV/Excel file headers and suggest column mappings."""
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Filename is required.")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise HTTPException(status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE, detail="Only .csv and .xlsx files are supported.")

    # Read headers
    headers: list[str] = []
    if suffix == ".csv":
        contents = await file.read()
        lines = contents.decode("utf-8").splitlines()
        if lines:
            reader = csv.reader([lines[0]])
            headers = next(reader, [])
    else:
        workbook = load_workbook(filename=BytesIO(await file.read()), read_only=True, data_only=True)
        worksheet = workbook.active
        first_row = next(worksheet.iter_rows(values_only=True), None)
        if first_row:
            headers = [str(cell) if cell is not None else "" for cell in first_row]

    # Suggest mappings based on common patterns
    suggested_mappings = _suggest_column_mappings(headers)

    return {
        "fileName": file.filename,
        "detectedColumns": headers,
        "suggestedMappings": suggested_mappings,
        "requiredFields": [
            {"field": "customer_id", "description": "Unique customer identifier", "required": True},
            {"field": "latitude", "description": "Latitude coordinate", "required": True},
            {"field": "longitude", "description": "Longitude coordinate", "required": True},
            {"field": "customer_name", "description": "Customer name", "required": False},
            {"field": "city", "description": "City or area", "required": False},
            {"field": "zone", "description": "Existing zone assignment", "required": False},
            {"field": "agent_id", "description": "Agent/Sales rep ID", "required": False},
            {"field": "agent_name", "description": "Agent/Sales rep name", "required": False},
            {"field": "status", "description": "Customer status", "required": False},
        ],
    }


def _suggest_column_mappings(headers: list[str]) -> dict[str, str]:
    """Auto-suggest column mappings based on header names."""
    mappings: dict[str, str] = {}
    normalized_headers = {h.lower().strip().replace(" ", "_").replace("-", "_"): h for h in headers}

    # Define patterns for each required field
    patterns = {
        "customer_id": ["customer_id", "customerid", "cusid", "id", "cust_id", "customer_code", "code", "account"],
        "latitude": ["latitude", "lat", "y", "coord_y"],
        "longitude": ["longitude", "lon", "lng", "long", "x", "coord_x"],
        "customer_name": ["customer_name", "customername", "cusname", "name", "cust_name", "account_name"],
        "city": ["city", "area", "region", "location", "town"],
        "zone": ["zone", "zone_id", "territory", "area_code", "sales_zone"],
        "agent_id": ["agent_id", "agentid", "agent_code", "sales_rep_id", "salesrep_id", "rep_id"],
        "agent_name": ["agent_name", "agentname", "sales_rep_name", "salesrep_name", "rep_name", "sales_rep"],
        "status": ["status", "state", "active", "customer_status"],
    }

    for field, pattern_list in patterns.items():
        for pattern in pattern_list:
            if pattern in normalized_headers:
                mappings[field] = normalized_headers[pattern]
                break

    return mappings


@router.post("/upload", response_model=CustomerStatsResponse, status_code=status.HTTP_201_CREATED)
async def upload_customer_dataset(
    file: UploadFile = File(...),
    mappings: str = Form(None),
    filter_columns: str = Form(None),
) -> CustomerStatsResponse:
    """Upload customer dataset with optional column mappings.

    Args:
        file: CSV or Excel file with customer data
        mappings: JSON string of column mappings (e.g., {"customer_id": "CusId", "latitude": "Lat"})
        filter_columns: JSON string array of columns to use as filters (e.g., ["City", "Zone"])
    """
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Filename is required.")

    suffix = Path(file.filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise HTTPException(status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE, detail="Only .csv and .xlsx files are supported.")

    # Parse mappings if provided
    column_mappings: dict[str, str] = {}
    if mappings:
        try:
            column_mappings = json.loads(mappings)
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid mappings JSON")

    # Parse filter columns if provided
    filter_cols: list[str] = []
    if filter_columns:
        try:
            filter_cols = json.loads(filter_columns)
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid filter_columns JSON")

    uploads_dir = settings.data_root / "uploads"
    uploads_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")
    destination = uploads_dir / f"customers_{timestamp}.csv"

    # Read file content
    if suffix == ".csv":
        contents = await file.read()
        lines = contents.decode("utf-8").splitlines()
        reader = csv.DictReader(lines)
        rows = list(reader)
    else:
        workbook = load_workbook(filename=BytesIO(await file.read()), read_only=True, data_only=True)
        worksheet = workbook.active
        headers = [str(cell) if cell is not None else "" for cell in next(worksheet.iter_rows(values_only=True), [])]
        rows = []
        for row_values in worksheet.iter_rows(values_only=True, min_row=2):
            row_dict = {headers[i]: ("" if cell is None else cell) for i, cell in enumerate(row_values) if i < len(headers)}
            rows.append(row_dict)

    # Apply column mappings if provided
    if column_mappings:
        # Remap columns: create new rows with standard column names
        remapped_rows = []
        for row in rows:
            remapped_row = {}
            # Copy mapped columns
            for standard_field, source_column in column_mappings.items():
                if source_column in row:
                    remapped_row[standard_field] = row[source_column]
            # Copy unmapped columns (preserve original data)
            for col, value in row.items():
                if col not in column_mappings.values():
                    remapped_row[col] = value
            remapped_rows.append(remapped_row)
        rows = remapped_rows

    # Write to CSV
    if rows:
        all_columns = list(dict.fromkeys([col for row in rows for col in row.keys()]))
        with destination.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=all_columns)
            writer.writeheader()
            writer.writerows(rows)
    else:
        # Empty file - just write headers
        with destination.open("w", newline="", encoding="utf-8") as handle:
            handle.write("")

    set_active_customer_file(destination.resolve())

    # Store filter column metadata
    if filter_cols:
        filter_meta_path = settings.data_root / "filter_metadata.json"
        try:
            with filter_meta_path.open("w", encoding="utf-8") as f:
                json.dump({"filter_columns": filter_cols, "updated_at": datetime.now(timezone.utc).isoformat()}, f, indent=2)
        except Exception:
            pass  # Don't fail upload if metadata storage fails

    stats = compute_customer_stats()
    return CustomerStatsResponse(**stats)


@router.get("/filters")
def get_filter_metadata() -> dict:
    """Get available filter columns from metadata."""
    filter_meta_path = settings.data_root / "filter_metadata.json"

    if not filter_meta_path.exists():
        return {"filter_columns": [], "updated_at": None}

    try:
        with filter_meta_path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"filter_columns": [], "updated_at": None}


@router.get("/filter-values/{column_name}")
def get_column_values(column_name: str, limit: int = Query(100, ge=1, le=1000)) -> dict:
    """Get unique values for a specific column to use in filter dropdowns."""
    try:
        customers = load_customers()
    except FileNotFoundError:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No customer data loaded")

    # Normalize column name (remove underscores, lowercase)
    normalized_col = column_name.lower().replace("_", "").replace("-", "")

    # Map standard field names to customer attributes
    field_mapping = {
        "city": "city",
        "zone": "zone",
        "zoneid": "zone",
        "agentid": "agent_id",
        "agentname": "agent_name",
        "status": "status",
        "area": "area",
        "region": "region",
    }

    attr_name = field_mapping.get(normalized_col)
    if not attr_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Column '{column_name}' not supported for filtering")

    # Collect unique values
    values_set = set()
    for customer in customers:
        value = getattr(customer, attr_name, None)
        if value and isinstance(value, str):
            value = value.strip()
            if value:
                values_set.add(value)

    # Sort and limit
    sorted_values = sorted(values_set)[:limit]

    return {
        "column": column_name,
        "values": sorted_values,
        "total_unique": len(values_set),
        "returned": len(sorted_values)
    }


@router.get("/current-file-info")
def get_current_file_info() -> dict:
    """Get information about the currently loaded customer file for re-mapping."""
    try:
        # Read current file headers
        file_path = settings.customer_file
        if not file_path.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No customer file loaded")

        # Read CSV headers
        headers: list[str] = []
        with file_path.open("r", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader, [])

        if not headers:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File has no headers")

        # Get current filter columns
        filter_meta_path = settings.data_root / "filter_metadata.json"
        current_filter_columns: list[str] = []
        if filter_meta_path.exists():
            try:
                with filter_meta_path.open("r", encoding="utf-8") as f:
                    metadata = json.load(f)
                    current_filter_columns = metadata.get("filter_columns", [])
            except Exception:
                pass

        # Suggest mappings
        suggested_mappings = _suggest_column_mappings(headers)

        return {
            "fileName": file_path.name,
            "detectedColumns": headers,
            "suggestedMappings": suggested_mappings,
            "currentFilterColumns": current_filter_columns,
            "requiredFields": [
                {"field": "customer_id", "description": "Unique customer identifier", "required": True},
                {"field": "latitude", "description": "Latitude coordinate", "required": True},
                {"field": "longitude", "description": "Longitude coordinate", "required": True},
                {"field": "customer_name", "description": "Customer name", "required": False},
                {"field": "city", "description": "City or area", "required": False},
                {"field": "zone", "description": "Existing zone assignment", "required": False},
                {"field": "agent_id", "description": "Agent/Sales rep ID", "required": False},
                {"field": "agent_name", "description": "Agent/Sales rep name", "required": False},
                {"field": "status", "description": "Customer status", "required": False},
            ],
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@router.put("/reprocess-mappings", response_model=CustomerStatsResponse, status_code=status.HTTP_200_OK)
def reprocess_current_file_mappings(
    mappings: str = Form(None),
    filter_columns: str = Form(None),
) -> CustomerStatsResponse:
    """Re-process the current customer file with new column mappings.

    Args:
        mappings: JSON string of column mappings (e.g., {"customer_id": "CusId", "latitude": "Lat"})
        filter_columns: JSON string array of columns to use as filters (e.g., ["City", "Zone"])
    """
    try:
        # Get current file path
        file_path = settings.customer_file
        if not file_path.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No customer file loaded")

        # Parse mappings if provided
        column_mappings: dict[str, str] = {}
        if mappings:
            try:
                column_mappings = json.loads(mappings)
            except json.JSONDecodeError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid mappings JSON")

        # Parse filter columns if provided
        filter_cols: list[str] = []
        if filter_columns:
            try:
                filter_cols = json.loads(filter_columns)
            except json.JSONDecodeError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid filter_columns JSON")

        # Read current CSV file
        with file_path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)

        # Apply column mappings if provided
        if column_mappings:
            remapped_rows = []
            for row in rows:
                remapped_row = {}
                # Copy mapped columns
                for standard_field, source_column in column_mappings.items():
                    if source_column in row:
                        remapped_row[standard_field] = row[source_column]
                # Copy unmapped columns (preserve original data)
                for col, value in row.items():
                    if col not in column_mappings.values():
                        remapped_row[col] = value
                remapped_rows.append(remapped_row)
            rows = remapped_rows

        # Re-write CSV with new mappings
        if rows:
            all_columns = list(dict.fromkeys([col for row in rows for col in row.keys()]))
            with file_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=all_columns)
                writer.writeheader()
                writer.writerows(rows)

        # Update filter column metadata
        if filter_cols:
            filter_meta_path = settings.data_root / "filter_metadata.json"
            try:
                with filter_meta_path.open("w", encoding="utf-8") as f:
                    json.dump({"filter_columns": filter_cols, "updated_at": datetime.now(timezone.utc).isoformat()}, f, indent=2)
            except Exception:
                pass  # Don't fail if metadata storage fails

        # Return updated stats
        stats = compute_customer_stats()
        return CustomerStatsResponse(**stats)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))
